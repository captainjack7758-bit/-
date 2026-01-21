import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import json
import pandas as pd
import os
import re
from openai import OpenAI
from copy import copy  # <---【必须添加这一行】用于复制样式
import threading
import openpyxl
from openpyxl import load_workbook
CONFIG_FILE = "config.json"

class TicketExtractorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("票务信息智能提取助手")
        self.root.geometry("1000x800")
      
        style = ttk.Style()
        style.theme_use('clam')
        style.configure("TLabel", font=("Microsoft YaHei", 10))
        style.configure("TButton", font=("Microsoft YaHei", 10))
        style.configure("Treeview.Heading", font=("Microsoft YaHei", 10, "bold"))
        style.configure("Treeview", font=("Microsoft YaHei", 10), rowheight=25)

        self.preview_data = []

        # 1. 设置区域
        settings_frame = ttk.LabelFrame(root, text="LLM 设置 (OpenAI 格式)", padding=10)
        settings_frame.pack(fill="x", padx=10, pady=5)

        grid_opts = {'padx': 5, 'pady': 5, 'sticky': 'w'}
      
        ttk.Label(settings_frame, text="Base URL:").grid(row=0, column=0, **grid_opts)
        self.entry_base_url = ttk.Entry(settings_frame, width=40)
        self.entry_base_url.grid(row=0, column=1, **grid_opts)

        ttk.Label(settings_frame, text="API Key:").grid(row=0, column=2, **grid_opts)
        self.entry_api_key = ttk.Entry(settings_frame, width=40, show="*")
        self.entry_api_key.grid(row=0, column=3, **grid_opts)

        ttk.Label(settings_frame, text="Model Name:").grid(row=1, column=0, **grid_opts)
        self.entry_model = ttk.Entry(settings_frame, width=40)
        self.entry_model.grid(row=1, column=1, **grid_opts)

        # 缩写替换设置
        abbrev_frame = ttk.LabelFrame(root, text="说明者缩写替换", padding=10)
        abbrev_frame.pack(fill="x", padx=10, pady=5)
        ttk.Label(abbrev_frame, text="格式: 原名1,缩写1,原名2,缩写2").pack(side="left", padx=5)
        self.entry_abbreviations = ttk.Entry(abbrev_frame, width=60)
        self.entry_abbreviations.pack(side="left", padx=5, fill="x", expand=True)

        # 加载配置
        self.load_config()

        # 绑定自动保存
        for entry in [self.entry_base_url, self.entry_api_key, self.entry_model, self.entry_abbreviations]:
            entry.bind("<FocusOut>", lambda e: self.save_config())

        # 2. Excel 文件选择
        file_frame = ttk.LabelFrame(root, text="Excel 文件设置", padding=10)
        file_frame.pack(fill="x", padx=10, pady=5)

        ttk.Label(file_frame, text="目标文件路径:").pack(side="left", padx=5)
        self.entry_file_path = ttk.Entry(file_frame, width=60)
        self.entry_file_path.pack(side="left", padx=5, fill="x", expand=True)
        ttk.Button(file_frame, text="选择/新建", command=self.browse_file).pack(side="left", padx=5)

        # 3. 文本输入区域
        input_frame = ttk.LabelFrame(root, text="聊天记录粘贴区", padding=10)
        input_frame.pack(fill="both", expand=True, padx=10, pady=5)

        self.text_input = scrolledtext.ScrolledText(input_frame, height=10, font=("Consolas", 10))
        self.text_input.pack(fill="both", expand=True)

        # 按钮区
        btn_frame = ttk.Frame(root)
        btn_frame.pack(fill="x", padx=10, pady=5)
      
        self.btn_extract = ttk.Button(btn_frame, text="开始提取 (AI解析)", command=self.run_extraction_thread)
        self.btn_extract.pack(side="left", padx=5)
      
        ttk.Label(btn_frame, text=" -> 检查预览 -> ").pack(side="left")
        self.btn_save = ttk.Button(btn_frame, text="保存到 Excel", command=self.save_to_excel, state="disabled")
        self.btn_save.pack(side="left", padx=5)

        self.status_var = tk.StringVar()
        self.status_var.set("就绪")
        ttk.Label(btn_frame, textvariable=self.status_var, foreground="blue").pack(side="right", padx=10)

        # 4. 预览区域
        preview_frame = ttk.LabelFrame(root, text="数据预览", padding=10)
        preview_frame.pack(fill="both", expand=True, padx=10, pady=10)

        columns = ("说明", "演出时间", "购买票价", "观演者姓名", "观演者身份证号码", "观演者手机号码", "购买数量", "卖价")
        self.tree = ttk.Treeview(preview_frame, columns=columns, show="headings")
        col_widths = [80, 100, 80, 80, 150, 100, 60, 80]
        for col, width in zip(columns, col_widths):
            self.tree.heading(col, text=col)
            self.tree.column(col, width=width, anchor="center")

        scrollbar = ttk.Scrollbar(preview_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscroll=scrollbar.set)
      
        self.tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

    def load_config(self):
        defaults = {
            "base_url": "https://api.openai.com/v1",
            "api_key": "",
            "model": "gpt-4o",
            "abbreviations": ""
        }
        config = defaults.copy()
        if os.path.exists(CONFIG_FILE):
            try:
                with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                    config.update(json.load(f))
            except Exception:
                pass
      
        self.entry_base_url.insert(0, config.get("base_url", defaults["base_url"]))
        self.entry_api_key.insert(0, config.get("api_key", ""))
        self.entry_model.insert(0, config.get("model", defaults["model"]))
        self.entry_abbreviations.insert(0, config.get("abbreviations", ""))

    def save_config(self):
        config = {
            "base_url": self.entry_base_url.get().strip(),
            "api_key": self.entry_api_key.get().strip(),
            "model": self.entry_model.get().strip(),
            "abbreviations": self.entry_abbreviations.get().strip()
        }
        try:
            with open(CONFIG_FILE, "w", encoding="utf-8") as f:
                json.dump(config, f, indent=2, ensure_ascii=False)
        except Exception:
            pass

    def browse_file(self):
        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile="票务统计.xlsx"
        )
        if filename:
            self.entry_file_path.delete(0, tk.END)
            self.entry_file_path.insert(0, filename)

    def run_extraction_thread(self):
        threading.Thread(target=self.extract_info, daemon=True).start()

    def convert_to_string(self, value):
        """将值转换为字符串，处理科学计数法和浮点数"""
        if value is None or (isinstance(value, float) and pd.isna(value)):
            return ""
        if isinstance(value, float):
            # 处理科学计数法，转为整数字符串
            return str(int(value))
        if isinstance(value, int):
            return str(value)
        return str(value).strip()

    def extract_info(self):
        text_content = self.text_input.get("1.0", tk.END).strip()
        if not text_content:
            messagebox.showwarning("提示", "请先粘贴聊天记录！")
            return
      
        base_url = self.entry_base_url.get().strip()
        api_key = self.entry_api_key.get().strip()
        model = self.entry_model.get().strip()

        if not api_key:
            messagebox.showerror("错误", "请填写 API Key")
            return

        self.btn_extract.config(state="disabled")
        self.status_var.set("正在请求 AI 进行分析，请稍候...")
        self.root.update_idletasks()

        try:
            client = OpenAI(api_key=api_key, base_url=base_url)
          
            system_prompt = """
你是一个票务数据提取助手。请将聊天记录转换为 JSON 数组。

### 字段定义
- "说明": 商家名（发消息的人）
- "演出时间": 格式统一为 "X月X日"（如 "1月16日"），根据聊天记录推断月份
- "购买票价": 票面原价（数字）
- "观演者姓名": 姓名（字符串）
- "观演者身份证号码": 身份证号（必须是字符串格式，18位）
- "观演者手机号码": 手机号（必须是字符串格式，11位）
- "卖价": 实际成交单价（数字）
- "购买数量": 购买数量量（数字，默认1）

### 演出时间处理规则
当前日期：1月19日
- "18号" → "1月18日"
- "1.16" → "1月16日"
- "25号" → "1月25日"（本月未过的日期）
- "2号" → "2月2日"（已过的日期推断为下月）

### 重要：身份证和手机号必须是字符串
身份证号码和手机号码必须用引号包裹，作为字符串输出，例如：
"观演者身份证号码": "420101199001011234"
"观演者手机号码": "13800000000"

### 示例

**示例1：单人**
输入：
海洋微信不收款转支付宝 
1.16 武汉梓渝 1280
张三 420101199001011234 13800000000
海洋微信不收款转支付宝 
1500

输出：
[{"说明": "海洋微信不收款转支付宝", "演出时间": "1月16日", "购买票价": 1280, "观演者姓名": "张三", "观演者身份证号码": "420101199001011234", "观演者手机号码": "13800000000", "购买数量": 1, "卖价": 1500}]

**示例2：回溯定价**
输入：
海洋微信不收款转支付宝 
田甜 421126199309090040
18号武汉梓渝 1280
17324862256
海洋微信不收款转支付宝 
25号武汉梓渝 980
黄诗涵 220523200702083226
17620370913
海洋微信不收款转支付宝 
1700

输出：
[{"说明": "海洋微信不收款转支付宝", "演出时间": "1月18日", "购买票价": 1280, "观演者姓名": "田甜", "观演者身份证号码": "421126199309090040", "观演者手机号码": "17324862256", "购买数量": 1, "卖价": 1700},
  {"说明": "海洋微信不收款转支付宝", "演出时间": "1月25日", "购买票价": 980, "观演者姓名": "黄诗涵", "观演者身份证号码": "220523200702083226", "观演者手机号码": "17620370913", "购买数量": 1, "卖价": 1700}
]

**示例3：乘法算式**
输入：
海洋微信不收款转支付宝 
16号武汉 1280
张三 420101199001011234
李四 420101199001011235
13800000000
海洋微信不收款转支付宝 
1500*2=3000

输出：
[
  {"说明": "海洋微信不收款转支付宝", "演出时间": "1月16日", "购买票价": 1280, "观演者姓名": "张三", "观演者身份证号码": "420101199001011234", "观演者手机号码": "13800000000", "购买数量": 1, "卖价": 1500},
  {"说明": "海洋微信不收款转支付宝", "演出时间": "1月16日", "购买票价": 1280, "观演者姓名": "李四", "观演者身份证号码": "420101199001011235", "观演者手机号码": "13800000000", "购买数量": 1, "卖价": 1500}
]

**示例4：加法算式**
输入：
海洋微信不收款转支付宝 
16号武汉 980
赵六 420101199001011236 13800000001
海洋微信不收款转支付宝 
20号武汉 1280
孙七 420101199001011237 13800000002
海洋微信不收款转支付宝 
1700+1800

输出：
[
  {"说明": "海洋微信不收款转支付宝", "演出时间": "1月16日", "购买票价": 980, "观演者姓名": "赵六", "观演者身份证号码": "420101199001011236", "观演者手机号码": "13800000001", "购买数量": 1, "卖价": 1700},
  {"说明": "海洋微信不收款转支付宝", "演出时间": "1月20日", "购买票价": 1280, "观演者姓名": "孙七", "观演者身份证号码": "420101199001011237", "观演者手机号码": "13800000002", "购买数量": 1, "卖价": 1800}
]

### 规则
1. 商家名单独出现表示新订单开始，最后单独出现的数字/算式是卖价
2. 卖价回溯应用到前面所有未定价的订单
3. "1500*2" 单价为1500；"1700+1800" 按顺序分配
4. 演出时间统一为 "X月X日" 格式，只有日期时根据当前日期推断月份
5. 数字字段（购买票价、购买数量、卖价）必须是数字类型
6. 身份证号码和手机号码必须是字符串类型（用引号包裹）
7. 输出纯 JSON 数组，不要 Markdown

            """

            response = client.chat.completions.create(
                model=model,
                messages=[
                    {"role": "system", "content": system_prompt}, 
                    {"role": "user", "content": text_content}
                ],
                temperature=0.7,
                extra_body={"reasoning_effort": "minimal"},
                response_format={"type":"json_object"},
                max_tokens=32000
            )

            result_text = response.choices[0].message.content.strip()
            if result_text.startswith("```json"):
                result_text = result_text[7:]
            if result_text.endswith("```"):
                result_text = result_text[:-3]
            with open('result.text','w') as f:
                f.write(result_text)
            data = json.loads(result_text)
            
            # 如果返回的是字典且包含数组，提取数组
            if isinstance(data, dict):
                for key in data:
                    if isinstance(data[key], list):
                        data = data[key]
                        break
            
            # 确保身份证和手机号是字符串
            for row in data:
                if "观演者身份证号码" in row:
                    row["观演者身份证号码"] = self.convert_to_string(row["观演者身份证号码"])
                if "观演者手机号码" in row:
                    row["观演者手机号码"] = self.convert_to_string(row["观演者手机号码"])
            
            self.root.after(0, self.update_preview, data)

        except Exception as e:
            self.root.after(0, lambda: messagebox.showerror("错误", f"提取失败: {str(e)}"))
            self.root.after(0, lambda: self.status_var.set("提取失败"))
        finally:
            self.root.after(0, lambda: self.btn_extract.config(state="normal"))

    def update_preview(self, data):
        self.preview_data = data
        for item in self.tree.get_children():
            self.tree.delete(item)
      
        for row in data:
            values = (
                row.get("说明", ""),
                row.get("演出时间", ""),
                row.get("购买票价", ""),
                row.get("观演者姓名", ""),
                row.get("观演者身份证号码", ""),
                row.get("观演者手机号码", ""),
                row.get("购买数量", 1),
                row.get("卖价", "")
            )
            self.tree.insert("", "end", values=values)
      
        self.btn_save.config(state="normal")
        self.status_var.set(f"提取成功！共 {len(data)} 条记录。请检查预览，确认无误后点击保存。")

    def save_to_excel(self):
        file_path = self.entry_file_path.get().strip()
        if not file_path:
            messagebox.showwarning("提示", "请选择保存路径")
            return
        
        if not self.preview_data:
            messagebox.showwarning("提示", "没有可保存的数据")
            return

        try:
            # --- 1. 数据准备 ---
            new_df = pd.DataFrame(self.preview_data)
            
            # 应用缩写替换
            abbrev_input = self.entry_abbreviations.get().strip()
            if abbrev_input:
                parts = [s.strip() for s in re.split(r'[,，]', abbrev_input)]
                if len(parts) >= 2:
                    abbrev_map = {parts[i]: parts[i+1] for i in range(0, len(parts)-1, 2)}
                    new_df["说明"] = new_df["说明"].str.strip().replace(abbrev_map)
            
            target_columns = ["说明", "演出时间", "购买票价", "观演者姓名", "观演者身份证号码", "观演者手机号码", "购买数量", "卖价"]
            
            # 补齐列
            for col in target_columns:
                if col not in new_df.columns:
                    new_df[col] = "" if col not in ["购买票价", "购买数量", "卖价"] else None
            
            new_df = new_df[target_columns]
            
            # 格式转换
            for col in ["观演者身份证号码", "观演者手机号码"]:
                new_df[col] = new_df[col].apply(self.convert_to_string)
            
            for col in ["购买票价", "购买数量", "卖价"]:
                new_df[col] = pd.to_numeric(new_df[col], errors="coerce")
            
            new_df["购买数量"] = new_df["购买数量"].fillna(1).astype(int)

            # --- 2. 查重逻辑 (新增核心部分) ---
            existing_ids = set()
            if os.path.exists(file_path):
                try:
                    # 读取现有文件中的身份证号，强制转为字符串并去空格
                    existing_df = pd.read_excel(file_path, sheet_name='信息', dtype={'观演者身份证号码': str})
                    if '观演者身份证号码' in existing_df.columns:
                        existing_ids = set(existing_df['观演者身份证号码'].dropna().astype(str).str.strip())
                except Exception:
                    # 如果读取失败（比如Sheet不存在），则认为没有重复数据
                    pass

            # 标记重复数据
            # 确保新数据的身份证号也是干净的字符串
            new_df['check_id'] = new_df['观演者身份证号码'].astype(str).str.strip()
            
            # 筛选出不在 existing_ids 中的数据
            initial_count = len(new_df)
            new_df = new_df[~new_df['check_id'].isin(existing_ids)]
            final_count = len(new_df)
            skipped_count = initial_count - final_count

            # 删除临时辅助列
            del new_df['check_id']

            if final_count == 0:
                messagebox.showinfo("提示", f"所有数据 ({skipped_count} 条) 均已存在于 Excel 中，未写入任何新数据。")
                return

            # --- 3. 写入 Excel (带样式克隆) ---
            
            if not os.path.exists(file_path):
                # 新文件直接写入
                with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                    new_df.to_excel(writer, sheet_name='信息', index=False)
                messagebox.showinfo("成功", f"新文件已创建：\n{file_path}\n成功写入: {final_count} 条")
                self.cleanup_ui()
                return

            # 追加写入
            wb = load_workbook(file_path)
            
            if '信息' in wb.sheetnames:
                ws = wb['信息']
            else:
                ws = wb.create_sheet('信息')
                ws.append(target_columns)

            start_row = ws.max_row + 1
            template_row_idx = ws.max_row if ws.max_row >= 2 else 1
            
            data_rows = new_df.values.tolist()

            for i, row_data in enumerate(data_rows):
                current_row = start_row + i
                for j, value in enumerate(row_data):
                    col_idx = j + 1
                    target_cell = ws.cell(row=current_row, column=col_idx)
                    
                    target_cell.value = value
                    
                    # 样式克隆
                    source_cell = ws.cell(row=template_row_idx, column=col_idx)
                    if source_cell.font: target_cell.font = copy(source_cell.font)
                    if source_cell.border: target_cell.border = copy(source_cell.border)
                    if source_cell.fill: target_cell.fill = copy(source_cell.fill)
                    if source_cell.alignment: target_cell.alignment = copy(source_cell.alignment)
                    target_cell.number_format = source_cell.number_format

                    # 强制文本格式
                    if col_idx == 5 or col_idx == 6:
                        target_cell.number_format = '@'

            wb.save(file_path)
            
            msg = f"处理完成！\n成功写入: {final_count} 条\n跳过重复: {skipped_count} 条"
            messagebox.showinfo("保存成功", msg)
            self.cleanup_ui()

        except Exception as e:
            messagebox.showerror("保存失败", f"写入 Excel 时出错:\n{str(e)}\n请检查文件是否被打开。")

    def cleanup_ui(self):
        """保存成功后的界面清理"""
        self.status_var.set("保存成功")
        self.text_input.delete("1.0", tk.END)
        for item in self.tree.get_children():
            self.tree.delete(item)
        self.preview_data = []
        self.btn_save.config(state="disabled")

if __name__ == "__main__":
    root = tk.Tk()
    app = TicketExtractorApp(root)
    root.mainloop()
